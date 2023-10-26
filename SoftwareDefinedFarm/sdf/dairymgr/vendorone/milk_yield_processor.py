# System packages
from time import time
from typing import Any
from pathlib import Path


# Local packages
from sdf.dairymgr.base.spreadsheet_utils import (get_value, ExcelReader)
from sdf.dairymgr.proto import dairymgr_pb2 as dairymgr 
from sdf.dairymgr.proto.generated.vendorone import vendor_one_pb2 as vone
from sdf.dairymgr.base.global_defs import (FLOOR_MSG_SIZE, CEILING_MSG_SIZE,
                                           VENDOR_ONE_ROW_COUNTER, Casts) 
from sdf.utils.universal_base_class import UniversalBase


# Third party packages
from openpyxl.cell.cell import TYPE_NULL
from openpyxl import load_workbook


__author__ = "Gloire Rubambiza"
__email__ = "gbr26@cornell.edu"
__credits__ = ["Gloire Rubambiza"]


# @brief: A class for processing milk yield data.
class MilkYieldReader(UniversalBase):

    def __init__(self):
        super().__init__()


    def read(self, file_path: str):
        """
           Open the excel file for reading before forwarding to server.
           :param file_path: A string.
           :rtype: A list of FarmPCMessage objects.
        """
        filename = Path(file_path).name
        process_start_timer = time()        

        # Flag to keep track of whether file is split into different messages.
        fragmented_big_file = False

        # The list of and counter of messages to be put in the queue.
        split_messages = []
        message_chunk_counter = 0

        # The interface to an excel reader for the file.
        reader = ExcelReader(file_path)

        next_milk_yield_batch = vone.PerFileMilkData() 

	# Determines the appropriate number of rows before checking size.
        counter = 0

        for row in range(2, reader.active_sheet.max_row+1):
            msg = next_milk_yield_batch.entries.add()
            msg.key.id = self.get_value(reader, row, 2, Casts.INT)
            msg.key.group_today = self.get_value(reader, row, 3, Casts.INT)
            msg.key.group_yesterday = self.get_value(reader, row, 4,
                                                     Casts.INT)
            value = self.get_value(reader, row, 5, Casts.BOOL)
            if value != None:
                msg.key.wrong_group = value 

            # Lactation Status
            msg.status.lactation = self.get_value(reader, row, 6,
                                                  Casts.INT)
            msg.status.dim = self.get_value(reader, row, 7,
                                                  Casts.INT)
            msg.status.status = self.get_value(reader, row, 8,
                                                  Casts.STR)

            # Milk Yield
            msg.m_yield.daily_yield = self.get_value(reader, row, 9,
                                                   Casts.FLOAT)
            msg.m_yield.yield_1 = self.get_value(reader, row, 10,
                                                   Casts.FLOAT)
            msg.m_yield.yield_2 = self.get_value(reader, row, 11,
                                                  Casts.FLOAT)
            msg.m_yield.yield_3 = self.get_value(reader, row, 12,
                                                   Casts.FLOAT)
            msg.m_yield.daily_avg_yield = self.get_value(reader, row, 13,
                                                   Casts.FLOAT)
            msg.m_yield.avg_yield_1 = self.get_value(reader, row, 14,
                                                   Casts.FLOAT)
            msg.m_yield.avg_yield_2 = self.get_value(reader, row, 15,
                                                   Casts.FLOAT)
            msg.m_yield.avg_yield_3 = self.get_value(reader, row, 16,
                                                   Casts.FLOAT)
            msg.m_yield.daily_yield_perc = self.get_value(reader, row, 17,
                                                   Casts.SINT)
            msg.m_yield.yield_dev_1 = self.get_value(reader, row, 18,
                                                   Casts.SINT)
            msg.m_yield.yield_dev_2 = self.get_value(reader, row, 19,
                                                   Casts.SINT)
            msg.m_yield.yield_dev_3 = self.get_value(reader, row, 20,
                                                   Casts.SINT)

            # FlowRateFat
            msg.flow_rate_fat.perc_milk_from_weight = self.get_value(reader,
                                                              row, 21,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.avg_yield_amt = self.get_value(reader,
                                                              row, 22,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.flow_rate_1 = self.get_value(reader,
                                                              row, 23,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.flow_rate_2 = self.get_value(reader,
                                                              row, 24,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.flow_rate_3 = self.get_value(reader,
                                                              row, 25,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.daily_fat = self.get_value(reader,
                                                          row, 26,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.daily_fat_perc = self.get_value(reader,
                                                              row, 27,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.fat_perc_1 = self.get_value(reader,
                                                              row, 28,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.fat_perc_2 = self.get_value(reader,
                                                              row, 29,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.fat_perc_3 = self.get_value(reader,
                                                              row, 30,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.avg_fat_1 = self.get_value(reader,
                                                          row, 31,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.avg_fat_2 = self.get_value(reader,
                                                          row, 32,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.avg_fat_3 = self.get_value(reader,
                                                              row, 33,
                                                        Casts.FLOAT)
            msg.flow_rate_fat.daily_avg_fat_perc = self.get_value(reader,
                                                                  row, 34,
                                                            Casts.FLOAT)
            msg.flow_rate_fat.daily_fat_dev_perc = self.get_value(reader,
                                                                  row, 35,
                                                            Casts.SINT)
            msg.flow_rate_fat.fat_dev_perc_1 = self.get_value(reader,
                                                              row, 36,
                                                            Casts.SINT)
            msg.flow_rate_fat.fat_dev_perc_2 = self.get_value(reader,
                                                              row, 37,
                                                            Casts.SINT)
            msg.flow_rate_fat.fat_dev_perc_3 = self.get_value(reader,
                                                              row, 38,
                                                            Casts.SINT)
            # Protein 
            msg.protein.daily_protein = self.get_value(reader, row, 39,
                                                        Casts.FLOAT)
            msg.protein.daily_protein_perc = self.get_value(reader, row, 40,
                                                            Casts.FLOAT)
            msg.protein.protein_perc_1 = self.get_value(reader, row, 41,
                                                        Casts.FLOAT)
            msg.protein.protein_perc_2 = self.get_value(reader, row, 42,
                                                        Casts.FLOAT)
            msg.protein.protein_perc_3 = self.get_value(reader, row, 43,
                                                        Casts.FLOAT)
            msg.protein.daily_avg_protein_perc = self.get_value(reader, row,
                                                                44,
                                                            Casts.FLOAT)
            msg.protein.avg_protein_1 = self.get_value(reader, row, 45,
                                                       Casts.FLOAT)
            msg.protein.avg_protein_2 = self.get_value(reader, row, 46,
                                                       Casts.FLOAT)
            msg.protein.avg_protein_3 = self.get_value(reader, row, 47,
                                                        Casts.FLOAT)
            msg.protein.daily_protein_dev_perc = self.get_value(reader, row,
                                                                48,
                                                            Casts.SINT)
            msg.protein.protein_dev_perc_1 = self.get_value(reader, row,
                                                            49,
                                                            Casts.SINT)
            msg.protein.protein_dev_perc_2 = self.get_value(reader, row,
                                                            50,
                                                            Casts.SINT)
            msg.protein.protein_dev_perc_3 = self.get_value(reader, row,
                                                            51,
                                                            Casts.SINT)

            # FatOverProtein
            msg.ft_o_protein.daily_fat_over_protein = self.get_value(
                                                                 reader, row,
                                                                 52,
                                                            Casts.FLOAT)
            msg.ft_o_protein.fat_over_protein_1 = self.get_value(
                                                                 reader, row,
                                                                 53,
                                                            Casts.FLOAT)
            msg.ft_o_protein.fat_over_protein_2 = self.get_value(
                                                                 reader, row,
                                                                 54,
                                                            Casts.FLOAT)
            msg.ft_o_protein.fat_over_protein_3 = self.get_value(
                                                                 reader, row,
                                                                 55,
                                                            Casts.FLOAT)
            msg.ft_o_protein.daily_avg_fat_over_protein = self.get_value(
                                                                 reader, row,
                                                                 56,
                                                            Casts.FLOAT)
            msg.ft_o_protein.avg_fat_over_protein_1 = self.get_value(
                                                                 reader, row,
                                                                 57,
                                                            Casts.FLOAT)
            msg.ft_o_protein.avg_fat_over_protein_2 = self.get_value(
                                                                 reader, row,
                                                                 58,
                                                            Casts.FLOAT)
            msg.ft_o_protein.avg_fat_over_protein_3 = self.get_value(
                                                                 reader, row,
                                                                 59,
                                                            Casts.FLOAT)
            msg.ft_o_protein.daily_fat_over_protein_dev_perc = self.get_value(
                                                                 reader, row,
                                                                 60,
                                                            Casts.SINT)
            msg.ft_o_protein.fat_over_protein_dev_perc_1 = self.get_value(
                                                                 reader, row,
                                                                 61,
                                                            Casts.SINT)
            msg.ft_o_protein.fat_over_protein_dev_perc_2 = self.get_value(
                                                                 reader, row,
                                                                 62,
                                                            Casts.SINT)
            msg.ft_o_protein.fat_over_protein_dev_perc_3 = self.get_value(
                                                                 reader, row,
                                                                 63,
                                                            Casts.SINT)
            # LactationBlood 
            msg.lactation_blood.daily_lactation_perc = self.get_value(reader,
                                                               row, 64,
                                                            Casts.FLOAT)
            msg.lactation_blood.lactation_perc_1 = self.get_value(reader,
                                                               row, 65,
                                                            Casts.FLOAT)
            msg.lactation_blood.lactation_perc_2 = self.get_value(reader,
                                                               row, 66,
                                                            Casts.FLOAT)
            msg.lactation_blood.lactation_perc_3 = self.get_value(reader,
                                                               row, 67,
                                                            Casts.FLOAT)
            msg.lactation_blood.daily_avg_lactation_perc = self.get_value(reader,
                                                               row, 68,
                                                            Casts.FLOAT)
            msg.lactation_blood.avg_lactation_1 = self.get_value(reader,
                                                               row, 69,
                                                            Casts.FLOAT)
            msg.lactation_blood.avg_lactation_2 = self.get_value(reader,
                                                               row, 70,
                                                            Casts.FLOAT)
            msg.lactation_blood.avg_lactation_3 = self.get_value(reader,
                                                               row, 71,
                                                            Casts.FLOAT)
            msg.lactation_blood.daily_lactation_dev_perc = self.get_value(
                                                                reader, row, 72,
                                                              Casts.SINT)
            msg.lactation_blood.lactation_dev_perc_1 = self.get_value(reader,
                                                               row, 73,
                                                            Casts.SINT)

            msg.lactation_blood.lactation_dev_perc_2 = self.get_value(reader,
                                                               row, 74,
                                                            Casts.SINT)
            msg.lactation_blood.lactation_dev_perc_3 = self.get_value(reader,
                                                               row, 75,
                                                            Casts.SINT)
            msg.lactation_blood.daily_lactose = self.get_value(reader,
                                                               row, 76,
                                                            Casts.FLOAT)
            msg.lactation_blood.daily_blood_perc = self.get_value(reader,
                                                               row, 77,
                                                            Casts.FLOAT)
            msg.lactation_blood.blood_perc_1 = self.get_value(reader,
                                                               row, 78,
                                                            Casts.FLOAT)
            msg.lactation_blood.blood_perc_2 = self.get_value(reader,
                                                               row, 79,
                                                            Casts.FLOAT)
            msg.lactation_blood.blood_perc_3 = self.get_value(reader,
                                                               row, 80,
                                                            Casts.FLOAT)
            # SCC 
            # msg.scc.daily_scc = self.get_value(reader, row, 81,
            #                                        Casts.STR)
            # msg.scc.scc_1 = self.get_value(reader, row, 82,
            #                                        Casts.STR)
            # msg.scc.scc_2 = self.get_value(reader, row, 83,
            #                                        Casts.STR)
            # msg.scc.scc_3 = self.get_value(reader, row, 84,
            #                                        Casts.STR)
            # msg.scc.daily_avg_scc = self.get_value(reader, row, 85,
            #                                        Casts.STR)
            # msg.scc.avg_scc_1 = self.get_value(reader, row, 86,
            #                                        Casts.STR)
            # msg.scc.avg_scc_2 = self.get_value(reader, row, 87,
            #                                        Casts.STR)
            # msg.scc.avg_scc_3 = self.get_value(reader, row, 88,
            #                                        Casts.STR)

            # ProductionRateConduction 
            msg.prod_rate_cond.avg_prod_rate = self.get_value(reader, row,
                                                              81,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.prod_rate_1 = self.get_value(reader, row,
                                                            82,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.prod_rate_2 = self.get_value(reader, row,
                                                            83,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.prod_rate_3 = self.get_value(reader, row,
                                                            84,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.avg_prod_rate_1 = self.get_value(reader, row,
                                                                85,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.avg_prod_rate_2 = self.get_value(reader, row,
                                                                86,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.avg_prod_rate_3 = self.get_value(reader, row,
                                                                87,
                                                            Casts.FLOAT)
            msg.prod_rate_cond.prod_rate_dev_perc_1 = self.get_value(reader,
                                                                     row,
                                                                     88,
                                                            Casts.SINT)
            msg.prod_rate_cond.prod_rate_dev_perc_2 = self.get_value(reader,
                                                                     row,
                                                                     89,
                                                            Casts.SINT)
            msg.prod_rate_cond.prod_rate_dev_perc_3 = self.get_value(reader, row,
                                                                     90,
                                                                     Casts.SINT)
            msg.prod_rate_cond.avg_cond = self.get_value(reader, row,
                                                         91,
                                                         Casts.FLOAT)
            msg.prod_rate_cond.cond_1 = self.get_value(reader, row,
                                                       92,
                                                       Casts.FLOAT)
            msg.prod_rate_cond.cond_2 = self.get_value(reader, row,
                                                       93,
                                                       Casts.FLOAT)
            msg.prod_rate_cond.cond_3 = self.get_value(reader, row,
                                                       94,
                                                       Casts.FLOAT)
            msg.prod_rate_cond.avg_cond_1 = self.get_value(reader, row,
                                                           95,
                                                           Casts.FLOAT)
            msg.prod_rate_cond.avg_cond_2 = self.get_value(reader, row,
                                                           96,
                                                           Casts.FLOAT)
            msg.prod_rate_cond.avg_cond_3 = self.get_value(reader, row,
                                                           97,
                                                           Casts.FLOAT)
            msg.prod_rate_cond.cond_dev_perc_1 = self.get_value(reader, row,
                                                                98,
                                                            Casts.SINT)
            msg.prod_rate_cond.cond_dev_perc_2 = self.get_value(reader, row,
                                                               99,
                                                            Casts.SINT)
            msg.prod_rate_cond.cond_dev_perc_3 = self.get_value(reader, row,
                                                               100,
                                                               Casts.SINT)

            # AMT 
            msg.amt.avg_amt = self.get_value(reader, row, 101,
                                             Casts.FLOAT)
            msg.amt.amt_1 = self.get_value(reader, row, 102,
                                           Casts.FLOAT)
            msg.amt.amt_2 = self.get_value(reader, row, 103,
                                           Casts.FLOAT)
            msg.amt.amt_3 = self.get_value(reader, row, 104,
                                           Casts.FLOAT)
            msg.amt.avg_amt_1 = self.get_value(reader, row, 105,
                                               Casts.FLOAT)
            msg.amt.avg_amt_2 = self.get_value(reader, row, 106,
                                               Casts.FLOAT)
            msg.amt.avg_amt_3 = self.get_value(reader, row, 107,
                                               Casts.FLOAT)
            msg.amt.amt_dev_perc_1 = self.get_value(reader, row, 108,
                                                    Casts.SINT)
            msg.amt.amt_dev_perc_2 = self.get_value(reader, row, 109,
                                                    Casts.SINT)
            msg.amt.amt_dev_perc_3 = self.get_value(reader, row, 110,
                                                    Casts.SINT)
            # Activity 
            msg.vone_activity.avg_act = self.get_value(reader, row, 111,
                                                  Casts.INT)
            msg.vone_activity.activity_1 = self.get_value(reader, row, 112,
                                                     Casts.INT)
            msg.vone_activity.activity_2 = self.get_value(reader, row, 113,
                                                     Casts.INT)
            msg.vone_activity.activity_3 = self.get_value(reader, row, 114,
                                                     Casts.INT)
            msg.vone_activity.avg_act_1 = self.get_value(reader, row, 115,
                                                     Casts.INT)
            msg.vone_activity.avg_act_2 = self.get_value(reader, row, 116,
                                                    Casts.INT)
            msg.vone_activity.avg_act_3 = self.get_value(reader, row, 117,
                                                    Casts.INT)
            msg.vone_activity.activ_last_dev_perc = self.get_value(reader, row,
                                                              118,
                                                              Casts.SINT)
            msg.vone_activity.activ_dev_perc_1 = self.get_value(reader, row,
                                                           119,
                                                           Casts.SINT)
            msg.vone_activity.activ_dev_perc_2 = self.get_value(reader, row,
                                                           120,
                                                           Casts.SINT)
            msg.vone_activity.activ_dev_perc_3 = self.get_value(reader, row,
                                                           121,
                                                           Casts.SINT)
            # Rest 
            msg.rest.rest_time_1 = self.get_value(reader, row, 122,
                                                  Casts.INT)
            msg.rest.rest_time_2 = self.get_value(reader, row, 123,
                                                  Casts.INT)
            msg.rest.rest_time_3 = self.get_value(reader, row, 124,
                                                  Casts.INT)
            msg.rest.avg_rest_time_1 = self.get_value(reader, row, 125,
                                                      Casts.INT)
            msg.rest.avg_rest_time_2 = self.get_value(reader, row, 126,
                                                      Casts.INT)
            msg.rest.avg_rest_time_3 = self.get_value(reader, row, 127,
                                                      Casts.INT)
            msg.rest.rest_time_dev_1 = self.get_value(reader, row, 128,
                                                      Casts.SINT)
            msg.rest.rest_time_dev_2 = self.get_value(reader, row, 129,
                                                      Casts.SINT)
            msg.rest.rest_time_dev_3 = self.get_value(reader, row, 130,
                                                      Casts.SINT)
            msg.rest.rest_ratio_1 = self.get_value(reader, row, 131,
                                                   Casts.INT)
            msg.rest.rest_ratio_2 = self.get_value(reader, row, 132,
                                                   Casts.INT)
            msg.rest.rest_ratio_3 = self.get_value(reader, row, 133,
                                                   Casts.INT)
            msg.rest.avg_rest_ratio_1 = self.get_value(reader, row, 134,
                                                       Casts.INT)
            msg.rest.avg_rest_ratio_2 = self.get_value(reader, row, 135,
                                                       Casts.INT)
            msg.rest.avg_rest_ratio_3 = self.get_value(reader, row, 136,
                                                       Casts.INT)
            msg.rest.rest_ratio_dev_1 = self.get_value(reader, row, 137,
                                                       Casts.SINT)
            msg.rest.rest_ratio_dev_2 = self.get_value(reader, row, 138,
                                                       Casts.SINT)
            msg.rest.rest_ratio_dev_3 = self.get_value(reader, row, 139,
                                                       Casts.SINT)
            msg.rest.rest_over_bout_1 = self.get_value(reader, row, 140,
                                                       Casts.INT)
            msg.rest.rest_over_bout_2 = self.get_value(reader, row, 141,
                                                       Casts.INT)
            msg.rest.rest_over_bout_3 = self.get_value(reader, row, 142,
                                                       Casts.INT)
            msg.rest.avg_rest_over_bout_1 = self.get_value(reader, row, 143,
                                                           Casts.INT)
            msg.rest.avg_rest_over_bout_2 = self.get_value(reader, row, 144,
                                                           Casts.INT)
            msg.rest.avg_rest_over_bout_3 = self.get_value(reader, row, 145,
                                                           Casts.INT)
            msg.rest.rest_over_bout_dev_1 = self.get_value(reader, row, 146,
                                                           Casts.SINT)
            msg.rest.rest_over_bout_dev_2 = self.get_value(reader, row, 147,
                                                           Casts.SINT)
            msg.rest.rest_over_bout_dev_3 = self.get_value(reader, row, 148,
                                                           Casts.SINT)
            msg.rest.rest_over_restlessness_1 = self.get_value(reader,
                                                               row, 149,
                                                               Casts.FLOAT)
            msg.rest.rest_over_restlessness_2 = self.get_value(reader, row,
                                                               150,
                                                               Casts.FLOAT)
            msg.rest.rest_over_restlessness_3 = self.get_value(reader, row,
                                                               151,
                                                               Casts.FLOAT)
            msg.rest.avg_rest_over_restl_1 = self.get_value(reader, row,
                                                            152,
                                                            Casts.FLOAT)
            msg.rest.avg_rest_over_restl_2 = self.get_value(reader, row,
                                                            153,
                                                            Casts.FLOAT)
            msg.rest.avg_rest_over_restl_3 = self.get_value(reader, row,
                                                            154,
                                                            Casts.FLOAT)
            msg.rest.rest_over_restless_dev_1 = self.get_value(reader, row, 155,
                                                               Casts.SINT)
            msg.rest.rest_over_restless_dev_2 = self.get_value(reader, row, 156,
                                                               Casts.SINT)
            msg.rest.rest_over_restless_dev_3 = self.get_value(reader, row, 157,
                                                               Casts.SINT)

            # Weight 
            msg.weight_stats.weight = self.get_value(reader, row, 158, 
                                                     Casts.INT)
            msg.weight_stats.avg_weight = self.get_value(reader, row, 159, 
                                                         Casts.INT)
            msg.weight_stats.weight_1 = self.get_value(reader, row, 160, 
                                                       Casts.INT)
            msg.weight_stats.weight_2 = self.get_value(reader, row, 161, 
                                                       Casts.INT)
            msg.weight_stats.weight_3 = self.get_value(reader, row, 162, 
                                                       Casts.INT)
            msg.weight_stats.avg_weight_1 = self.get_value(reader, row, 163, 
                                                           Casts.INT)
            msg.weight_stats.avg_weight_2 = self.get_value(reader, row, 164, 
                                                           Casts.INT)
            msg.weight_stats.avg_weight_3 = self.get_value(reader, row, 165, 
                                                           Casts.INT)
            msg.weight_stats.weight_last_dev = self.get_value(reader, row, 166,
                                                              Casts.SINT)
            msg.weight_stats.weight_dev_1 = self.get_value(reader, row, 167, 
                                                           Casts.SINT)
            msg.weight_stats.weight_dev_2 = self.get_value(reader, row, 168, 
                                                           Casts.SINT)
            msg.weight_stats.weight_dev_3 = self.get_value(reader, row, 169, 
                                                           Casts.SINT)
            msg.weight_stats.weigh_in_cav = self.get_value(reader, row, 170, 
                                                           Casts.INT)
            msg.weight_stats.prev_1_weight = self.get_value(reader, row, 171, 
                                                            Casts.INT)
            msg.weight_stats.weight_in_breeding = self.get_value(reader, row,
                                                                 172,
                                                                 Casts.INT)
            msg.weight_stats.loss_of_weight = self.get_value(reader, row, 173, 
                                                             Casts.FLOAT)
            msg.weight_stats.perc_loss_of_weight = self.get_value(reader, row,
                                                                  174,
                                                                  Casts.INT)
            # Check message size and reinitialize as needed.
            if counter == VENDOR_ONE_ROW_COUNTER or row == reader.active_sheet.max_row:
                size = next_milk_yield_batch.ByteSize()
                if size < FLOOR_MSG_SIZE or size >= CEILING_MSG_SIZE:
                    print('%s: MSG size outside of window: %d' % (self.__class__.__name__, size) )
                if True:
                    # Create a new vendor one record regardless of window size.
                    # This is based on evidence from preliminary data pushes
                    # which shows that most messages are less than 8KB
                    entry = dairymgr.FarmPCMessage()
                    entry.vendorOne.filename = filename
                    entry.vendorOne.milk.CopyFrom(next_milk_yield_batch)
                    split_messages.append(entry)

                    # Prep next batch
                    next_milk_yield_batch = vone.PerFileMilkData()
                    fragmented_big_file = True
                    message_chunk_counter += 1

                # Reset the counter
                counter = 0
            else:
                counter += 1

        # Check if the file has been fragmented.
        if not(fragmented_big_file):
            entry = dairymgr.FarmPCMessage()
            entry.vendorOne.filename = filename
            entry.vendorOne.milk.CopyFrom(next_milk_yield_batch)
            split_messages.append(entry)

        processing_time = time() - process_start_timer
        self.log("Processing %s took %s mins" % (file_path,
                                           (processing_time / 60)))

        # Log the number of chunks for system debugging.
        self.log("%s split into %d chunks\n" % (file_path, message_chunk_counter))

        return split_messages 


    def get_value(self, reader, row_num, column, exp_type):
        """
           Get the value in a given row and column in excel sheet if it exists.
           :param reader: an ExcelReader object .
           :param row_num: An int.
           :param column: An int.
           :param exp_type: A Casts enum.
        """
        cell = reader.get_cell(row_num, column)
        value = None

        # Checking for TYPE_NULL is hard because it's the same as
        # the character for numeric types in openpyxl ('n')
        # Therefore we just check what null values may look like.
        null_vals = ["--", "", None, "N/A"]

        if cell.value in null_vals:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = 0
            elif exp_type == Casts.FLOAT:
                value = 0.0
            elif exp_type == Casts.STR:
                value = ""
            elif exp_type == Casts.BOOL:
                value = None
            else:
                raise ValueError("Unknown type (%s) at row: %d, col: %d!\n" \
                                  % (cell.cell.data_type, row_num, column))
        else:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = int(cell.value)
            elif exp_type == Casts.STR:
                value = str(cell.value)
            elif exp_type == Casts.FLOAT:
                value = float(cell.value)
            else: # Bool
                if cell.value == "No":
                    value = False
                else:
                    value = True
        return value
