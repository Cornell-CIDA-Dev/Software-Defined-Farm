# System packages
from time import time
from pathlib import Path


# Third party packages
import xlrd
from openpyxl.cell.cell import TYPE_NULL
from openpyxl import load_workbook


from sdf.dairymgr.proto import dairymgr_pb2 as dairymgr 
from sdf.dairymgr.proto.generated.vendorone import vendor_one_pb2 as vone 
from sdf.dairymgr.base.global_defs import Casts
from sdf.utils.universal_base_class import UniversalBase


__author__ = "Gloire Rubambiza"
__email__ = "gbr26@cornell.edu"
__credits__ = ["Gloire Rubambiza"]


# @brief: A class for processing milk yield data.
class MilkYieldReader:

    def process_file(self, file_path):
        """
           Open the excel file for reading before forwarding to server.
           :param file_path: A string.
        """
        filename = Path(file_path).name
        process_start_timer = time()        
        milk_yield_messages = []

        with xlrd.open_workbook(file_path, on_demand=True) as milk_workbook:
            sheet = milk_workbook.sheet_by_index(0)
            milk_file = vone.PerFileMilkData() 


            for row in range(1, sheet.nrows):
                msg = milk_file.entries.add()
                msg.key.id = self.get_value(sheet, row, 1, Casts.INT)
                msg.key.group_today = self.get_value(sheet, row, 2,
                                                     Casts.INT)
                msg.key.group_yesterday = self.get_value(sheet, row, 3,
                                                     Casts.INT)
                value = self.get_value(sheet, row, 4,
                                                     Casts.BOOL)
                if value != None:
                    msg.key.wrong_group = value 

                # Lactation Status
                msg.status.lactation = self.get_value(sheet, row, 5,
                                                      Casts.INT)
                msg.status.dim = self.get_value(sheet, row, 6,
                                                      Casts.INT)
                msg.status.status = self.get_value(sheet, row, 7,
                                                      Casts.STR)

                # Milk Yield
                msg.m_yield.daily_yield = self.get_value(sheet, row, 8,
                                                       Casts.FLOAT)
                msg.m_yield.yield_1 = self.get_value(sheet, row, 9,
                                                       Casts.FLOAT)
                msg.m_yield.yield_2 = self.get_value(sheet, row, 10,
                                                      Casts.FLOAT)
                msg.m_yield.yield_3 = self.get_value(sheet, row, 11,
                                                       Casts.FLOAT)
                msg.m_yield.daily_avg_yield = self.get_value(sheet, row, 12,
                                                       Casts.FLOAT)
                msg.m_yield.avg_yield_1 = self.get_value(sheet, row, 13,
                                                       Casts.FLOAT)
                msg.m_yield.avg_yield_2 = self.get_value(sheet, row, 14,
                                                       Casts.FLOAT)
                msg.m_yield.avg_yield_3 = self.get_value(sheet, row, 15,
                                                       Casts.FLOAT)
                msg.m_yield.daily_yield_perc = self.get_value(sheet, row, 16,
                                                       Casts.SINT)
                msg.m_yield.yield_dev_1 = self.get_value(sheet, row, 17,
                                                       Casts.SINT)
                msg.m_yield.yield_dev_2 = self.get_value(sheet, row, 18,
                                                       Casts.SINT)
                msg.m_yield.yield_dev_3 = self.get_value(sheet, row, 19,
                                                       Casts.SINT)

                # FlowRateFat
                msg.flow_rate_fat.perc_milk_from_weight = self.get_value(sheet,
                                                                  row, 20,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.avg_yield_amt = self.get_value(sheet,
                                                                  row, 21,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.flow_rate_1 = self.get_value(sheet,
                                                                  row, 22,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.flow_rate_2 = self.get_value(sheet,
                                                                  row, 23,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.flow_rate_3 = self.get_value(sheet,
                                                                  row, 24,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.daily_fat = self.get_value(sheet,
                                                                  row, 25,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.daily_fat_perc = self.get_value(sheet,
                                                                  row, 26,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.fat_perc_1 = self.get_value(sheet,
                                                                  row, 27,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.fat_perc_2 = self.get_value(sheet,
                                                                  row, 28,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.fat_perc_3 = self.get_value(sheet,
                                                                  row, 29,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.avg_fat_1 = self.get_value(sheet,
                                                                  row, 30,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.avg_fat_2 = self.get_value(sheet,
                                                                  row, 31,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.avg_fat_3 = self.get_value(sheet,
                                                                  row, 32,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.daily_avg_fat_perc = self.get_value(sheet,
                                                                  row, 33,
                                                            Casts.FLOAT)
                msg.flow_rate_fat.daily_fat_dev_perc = self.get_value(sheet,
                                                                  row, 34,
                                                            Casts.SINT)
                msg.flow_rate_fat.fat_dev_perc_1 = self.get_value(sheet,
                                                                  row, 35,
                                                            Casts.SINT)
                msg.flow_rate_fat.fat_dev_perc_2 = self.get_value(sheet,
                                                                  row, 36,
                                                            Casts.SINT)
                msg.flow_rate_fat.fat_dev_perc_3 = self.get_value(sheet,
                                                                  row, 37,
                                                            Casts.SINT)
                # Protein 
                msg.protein.daily_protein = self.get_value(sheet, row, 38,
                                                            Casts.FLOAT)
                msg.protein.daily_protein_perc = self.get_value(sheet, row, 39,
                                                            Casts.FLOAT)
                msg.protein.protein_perc_1 = self.get_value(sheet, row, 40,
                                                            Casts.FLOAT)
                msg.protein.protein_perc_2 = self.get_value(sheet, row, 41,
                                                            Casts.FLOAT)
                msg.protein.protein_perc_3 = self.get_value(sheet, row, 42,
                                                            Casts.FLOAT)
                msg.protein.daily_avg_protein_perc = self.get_value(sheet, row,
                                                                    43,
                                                            Casts.FLOAT)
                msg.protein.avg_protein_1 = self.get_value(sheet, row, 44,
                                                            Casts.FLOAT)
                msg.protein.avg_protein_2 = self.get_value(sheet, row, 45,
                                                            Casts.FLOAT)
                msg.protein.avg_protein_3 = self.get_value(sheet, row, 46,
                                                            Casts.FLOAT)
                msg.protein.daily_protein_dev_perc = self.get_value(sheet, row,
                                                                    47,
                                                            Casts.SINT)
                msg.protein.protein_dev_perc_1 = self.get_value(sheet, row,
                                                                    48,
                                                            Casts.SINT)
                msg.protein.protein_dev_perc_2 = self.get_value(sheet, row,
                                                                    49,
                                                            Casts.SINT)
                msg.protein.protein_dev_perc_3 = self.get_value(sheet, row,
                                                                    50,
                                                            Casts.SINT)

                # FatOverProtein
                msg.ft_o_protein.daily_fat_over_protein = self.get_value(
                                                                 sheet, row,
                                                                 51,
                                                            Casts.FLOAT)
                msg.ft_o_protein.fat_over_protein_1 = self.get_value(
                                                                 sheet, row,
                                                                 52,
                                                            Casts.FLOAT)
                msg.ft_o_protein.fat_over_protein_2 = self.get_value(
                                                                 sheet, row,
                                                                 53,
                                                            Casts.FLOAT)
                msg.ft_o_protein.fat_over_protein_3 = self.get_value(
                                                                 sheet, row,
                                                                 54,
                                                            Casts.FLOAT)
                msg.ft_o_protein.daily_avg_fat_over_protein = self.get_value(
                                                                 sheet, row,
                                                                 55,
                                                            Casts.FLOAT)
                msg.ft_o_protein.avg_fat_over_protein_1 = self.get_value(
                                                                 sheet, row,
                                                                 56,
                                                            Casts.FLOAT)
                msg.ft_o_protein.avg_fat_over_protein_2 = self.get_value(
                                                                 sheet, row,
                                                                 57,
                                                            Casts.FLOAT)
                msg.ft_o_protein.avg_fat_over_protein_3 = self.get_value(
                                                                 sheet, row,
                                                                 58,
                                                            Casts.FLOAT)
                msg.ft_o_protein.daily_fat_over_protein_dev_perc = self.get_value(
                                                                 sheet, row,
                                                                 59,
                                                            Casts.SINT)
                msg.ft_o_protein.fat_over_protein_dev_perc_1 = self.get_value(
                                                                 sheet, row,
                                                                 60,
                                                            Casts.SINT)
                msg.ft_o_protein.fat_over_protein_dev_perc_2 = self.get_value(
                                                                 sheet, row,
                                                                 61,
                                                            Casts.SINT)
                msg.ft_o_protein.fat_over_protein_dev_perc_3 = self.get_value(
                                                                 sheet, row,
                                                                 62,
                                                            Casts.SINT)
                # LactationBlood 
                msg.lactation_blood.daily_lactation_perc = self.get_value(sheet,
                                                               row, 63,
                                                            Casts.FLOAT)
                msg.lactation_blood.lactation_perc_1 = self.get_value(sheet,
                                                               row, 64,
                                                            Casts.FLOAT)
                msg.lactation_blood.lactation_perc_2 = self.get_value(sheet,
                                                               row, 65,
                                                            Casts.FLOAT)
                msg.lactation_blood.lactation_perc_3 = self.get_value(sheet,
                                                               row, 66,
                                                            Casts.FLOAT)
                msg.lactation_blood.daily_avg_lactation_perc = self.get_value(sheet,
                                                               row, 67,
                                                            Casts.FLOAT)
                msg.lactation_blood.avg_lactation_1 = self.get_value(sheet,
                                                               row, 68,
                                                            Casts.FLOAT)
                msg.lactation_blood.avg_lactation_2 = self.get_value(sheet,
                                                               row, 69,
                                                            Casts.FLOAT)
                msg.lactation_blood.avg_lactation_3 = self.get_value(sheet,
                                                               row, 70,
                                                            Casts.FLOAT)
                msg.lactation_blood.daily_lactation_dev_perc = self.get_value(
                                                                sheet, row, 71,
                                                              Casts.SINT)
                msg.lactation_blood.lactation_dev_perc_1 = self.get_value(sheet,
                                                               row, 72,
                                                            Casts.SINT)

                msg.lactation_blood.lactation_dev_perc_2 = self.get_value(sheet,
                                                               row, 73,
                                                            Casts.SINT)
                msg.lactation_blood.lactation_dev_perc_3 = self.get_value(sheet,
                                                               row, 74,
                                                            Casts.SINT)
                msg.lactation_blood.daily_lactose = self.get_value(sheet,
                                                               row, 75,
                                                            Casts.FLOAT)
                msg.lactation_blood.daily_blood_perc = self.get_value(sheet,
                                                               row, 76,
                                                            Casts.FLOAT)
                msg.lactation_blood.blood_perc_1 = self.get_value(sheet,
                                                               row, 77,
                                                            Casts.FLOAT)
                msg.lactation_blood.blood_perc_2 = self.get_value(sheet,
                                                               row, 78,
                                                            Casts.FLOAT)
                msg.lactation_blood.blood_perc_3 = self.get_value(sheet,
                                                               row, 79,
                                                            Casts.FLOAT)
                # SCC 
                msg.scc.daily_scc = self.get_value(sheet, row, 80,
                                                   Casts.STR)
                msg.scc.scc_1 = self.get_value(sheet, row, 81,
                                                   Casts.STR)
                msg.scc.scc_2 = self.get_value(sheet, row, 82,
                                                   Casts.STR)
                msg.scc.scc_3 = self.get_value(sheet, row, 83,
                                                   Casts.STR)
                msg.scc.daily_avg_scc = self.get_value(sheet, row, 84,
                                                   Casts.STR)
                msg.scc.avg_scc_1 = self.get_value(sheet, row, 85,
                                                   Casts.STR)
                msg.scc.avg_scc_2 = self.get_value(sheet, row, 86,
                                                   Casts.STR)
                msg.scc.avg_scc_3 = self.get_value(sheet, row, 87,
                                                   Casts.STR)

                # ProductionRateConduction 
                msg.prod_rate_cond.avg_prod_rate = self.get_value(sheet, row,
                                                                  88,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.prod_rate_1 = self.get_value(sheet, row,
                                                                  89,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.prod_rate_2 = self.get_value(sheet, row,
                                                                  90,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.prod_rate_3 = self.get_value(sheet, row,
                                                                  91,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_prod_rate_1 = self.get_value(sheet, row,
                                                                  92,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_prod_rate_2 = self.get_value(sheet, row,
                                                                  93,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_prod_rate_3 = self.get_value(sheet, row,
                                                                  94,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.prod_rate_dev_perc_1 = self.get_value(sheet,
                                                                         row,
                                                                         95,
                                                            Casts.SINT)
                msg.prod_rate_cond.prod_rate_dev_perc_2 = self.get_value(sheet,
                                                                         row,
                                                                         96,
                                                            Casts.SINT)
                msg.prod_rate_cond.prod_rate_dev_perc_3 = self.get_value(sheet, row,
                                                                  97,
                                                            Casts.SINT)
                msg.prod_rate_cond.avg_cond = self.get_value(sheet, row,
                                                                  98,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.cond_1 = self.get_value(sheet, row,
                                                                  99,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.cond_2 = self.get_value(sheet, row,
                                                                  100,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.cond_3 = self.get_value(sheet, row,
                                                                  101,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_cond_1 = self.get_value(sheet, row,
                                                                  102,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_cond_2 = self.get_value(sheet, row,
                                                                  103,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.avg_cond_3 = self.get_value(sheet, row,
                                                                  104,
                                                            Casts.FLOAT)
                msg.prod_rate_cond.cond_dev_perc_1 = self.get_value(sheet, row,
                                                                  105,
                                                            Casts.SINT)
                msg.prod_rate_cond.cond_dev_perc_2 = self.get_value(sheet, row,
                                                                  106,
                                                            Casts.SINT)
                msg.prod_rate_cond.cond_dev_perc_3 = self.get_value(sheet, row,
                                                                  107,
                                                            Casts.SINT)

                # AMT 
                msg.amt.avg_amt = self.get_value(sheet, row, 108,
                                                 Casts.FLOAT)
                msg.amt.amt_1 = self.get_value(sheet, row, 109,
                                                 Casts.FLOAT)
                msg.amt.amt_2 = self.get_value(sheet, row, 110,
                                                 Casts.FLOAT)
                msg.amt.amt_3 = self.get_value(sheet, row, 111,
                                                 Casts.FLOAT)
                msg.amt.avg_amt_1 = self.get_value(sheet, row, 112,
                                                 Casts.FLOAT)
                msg.amt.avg_amt_2 = self.get_value(sheet, row, 113,
                                                 Casts.FLOAT)
                msg.amt.avg_amt_3 = self.get_value(sheet, row, 114,
                                                 Casts.FLOAT)
                msg.amt.amt_dev_perc_1 = self.get_value(sheet, row, 115,
                                                 Casts.SINT)
                msg.amt.amt_dev_perc_2 = self.get_value(sheet, row, 116,
                                                 Casts.SINT)
                msg.amt.amt_dev_perc_3 = self.get_value(sheet, row, 117,
                                                 Casts.SINT)
                # Activity 
                msg.vone_activity.avg_act = self.get_value(sheet, row, 118,
                                                       Casts.INT)
                msg.vone_activity.activity_1 = self.get_value(sheet, row, 119,
                                                       Casts.INT)
                msg.vone_activity.activity_2 = self.get_value(sheet, row, 120,
                                                       Casts.INT)
                msg.vone_activity.activity_3 = self.get_value(sheet, row, 121,
                                                       Casts.INT)
                msg.vone_activity.avg_act_1 = self.get_value(sheet, row, 122,
                                                       Casts.INT)
                msg.vone_activity.avg_act_2 = self.get_value(sheet, row, 123,
                                                       Casts.INT)
                msg.vone_activity.avg_act_3 = self.get_value(sheet, row, 124,
                                                       Casts.INT)
                msg.vone_activity.activ_last_dev_perc = self.get_value(sheet,
                                                                           row,
                                                                           125,
                                                                    Casts.SINT)
                #msg.activity.activ_last_dev_perc = self.get_value(sheet, row,
                #                                                  126,
                #                                       Casts.SINT)
                msg.vone_activity.activ_dev_perc_1 = self.get_value(sheet, row,
                                                                  126,
                                                       Casts.SINT)
                msg.vone_activity.activ_dev_perc_2 = self.get_value(sheet, row,
                                                                  127,
                                                       Casts.SINT)
                msg.vone_activity.activ_dev_perc_3 = self.get_value(sheet, row,
                                                                  128,
                                                       Casts.SINT)
                # Rest 
                msg.rest.rest_time_1 = self.get_value(sheet, row, 129,
                                                 Casts.INT)
                msg.rest.rest_time_2 = self.get_value(sheet, row, 130,
                                                 Casts.INT)
                msg.rest.rest_time_3 = self.get_value(sheet, row, 131,
                                                 Casts.INT)
                msg.rest.avg_rest_time_1 = self.get_value(sheet, row, 132,
                                                 Casts.INT)
                msg.rest.avg_rest_time_2 = self.get_value(sheet, row, 133,
                                                 Casts.INT)
                msg.rest.avg_rest_time_3 = self.get_value(sheet, row, 134,
                                                 Casts.INT)
                msg.rest.rest_time_dev_1 = self.get_value(sheet, row, 135,
                                                 Casts.SINT)
                msg.rest.rest_time_dev_2 = self.get_value(sheet, row, 136,
                                                 Casts.SINT)
                msg.rest.rest_time_dev_3 = self.get_value(sheet, row, 137,
                                                 Casts.SINT)
                msg.rest.rest_ratio_1 = self.get_value(sheet, row, 138,
                                                 Casts.INT)
                msg.rest.rest_ratio_2 = self.get_value(sheet, row, 139,
                                                 Casts.INT)
                msg.rest.rest_ratio_3 = self.get_value(sheet, row, 140,
                                                 Casts.INT)
                msg.rest.avg_rest_ratio_1 = self.get_value(sheet, row, 141,
                                                 Casts.INT)
                msg.rest.avg_rest_ratio_2 = self.get_value(sheet, row, 142,
                                                 Casts.INT)
                msg.rest.avg_rest_ratio_3 = self.get_value(sheet, row, 143,
                                                 Casts.INT)
                msg.rest.rest_ratio_dev_1 = self.get_value(sheet, row, 144,
                                                 Casts.SINT)
                msg.rest.rest_ratio_dev_2 = self.get_value(sheet, row, 145,
                                                 Casts.SINT)
                msg.rest.rest_ratio_dev_3 = self.get_value(sheet, row, 146,
                                                 Casts.SINT)
                msg.rest.rest_over_bout_1 = self.get_value(sheet, row, 147,
                                                 Casts.INT)
                msg.rest.rest_over_bout_2 = self.get_value(sheet, row, 148,
                                                 Casts.INT)
                msg.rest.rest_over_bout_3 = self.get_value(sheet, row, 149,
                                                 Casts.INT)
                msg.rest.avg_rest_over_bout_1 = self.get_value(sheet, row, 150,
                                                 Casts.INT)
                msg.rest.avg_rest_over_bout_2 = self.get_value(sheet, row, 151,
                                                 Casts.INT)
                msg.rest.avg_rest_over_bout_3 = self.get_value(sheet, row, 152,
                                                 Casts.INT)
                msg.rest.rest_over_bout_dev_1 = self.get_value(sheet, row, 153,
                                                 Casts.SINT)
                msg.rest.rest_over_bout_dev_2 = self.get_value(sheet, row, 154,
                                                 Casts.SINT)
                msg.rest.rest_over_bout_dev_3 = self.get_value(sheet, row, 155,
                                                 Casts.SINT)
                msg.rest.rest_over_restlessness_1 = self.get_value(sheet,
                                                                      row, 156,
                                                          Casts.FLOAT)
                msg.rest.rest_over_restlessness_2 = self.get_value(sheet, row,
                                                                  157,
                                                          Casts.FLOAT)
                msg.rest.rest_over_restlessness_3 = self.get_value(sheet, row,
                                                                  158,
                                                          Casts.FLOAT)
                msg.rest.avg_rest_over_restl_1 = self.get_value(sheet, row,
                                                                  159,
                                                          Casts.FLOAT)
                msg.rest.avg_rest_over_restl_2 = self.get_value(sheet, row,
                                                                  160,
                                                          Casts.FLOAT)
                msg.rest.avg_rest_over_restl_3 = self.get_value(sheet, row,
                                                                  161,
                                                          Casts.FLOAT)
                msg.rest.rest_over_restless_dev_1 = self.get_value(sheet, row, 162,
                                                 Casts.SINT)
                msg.rest.rest_over_restless_dev_2 = self.get_value(sheet, row, 163,
                                                 Casts.SINT)
                msg.rest.rest_over_restless_dev_3 = self.get_value(sheet, row, 164,
                                                 Casts.SINT)

                # Weight 
                msg.weight_stats.weight = self.get_value(sheet, row, 165, 
                                                         Casts.INT)
                msg.weight_stats.avg_weight = self.get_value(sheet, row, 166, 
                                                         Casts.INT)
                msg.weight_stats.weight_1 = self.get_value(sheet, row, 167, 
                                                         Casts.INT)
                msg.weight_stats.weight_2 = self.get_value(sheet, row, 168, 
                                                         Casts.INT)
                msg.weight_stats.weight_3 = self.get_value(sheet, row, 169, 
                                                         Casts.INT)
                msg.weight_stats.avg_weight_1 = self.get_value(sheet, row, 170, 
                                                         Casts.INT)
                msg.weight_stats.avg_weight_2 = self.get_value(sheet, row, 171, 
                                                         Casts.INT)
                msg.weight_stats.avg_weight_3 = self.get_value(sheet, row, 172, 
                                                         Casts.INT)
                msg.weight_stats.weight_last_dev = self.get_value(sheet, row, 173, 
                                                         Casts.SINT)
                msg.weight_stats.weight_dev_1 = self.get_value(sheet, row, 174, 
                                                         Casts.SINT)
                msg.weight_stats.weight_dev_2 = self.get_value(sheet, row, 175, 
                                                         Casts.SINT)
                msg.weight_stats.weight_dev_3 = self.get_value(sheet, row, 176, 
                                                         Casts.SINT)
                msg.weight_stats.weigh_in_cav = self.get_value(sheet, row, 177, 
                                                         Casts.INT)
                msg.weight_stats.prev_1_weight = self.get_value(sheet, row, 178, 
                                                         Casts.INT)
                msg.weight_stats.loss_of_weight = self.get_value(sheet, row, 179, 
                                                         Casts.FLOAT)
                msg.weight_stats.perc_loss_of_weight = self.get_value(sheet, row, 180, 
                                                         Casts.INT)
                msg.weight_stats.weight_in_breeding = self.get_value(sheet, row, 181,
                                                         Casts.INT)
        entry = dairymgr.FarmPCMessage()
        entry.vendorOne.filename = filename
        entry.vendorOne.milk.CopyFrom(milk_file)
        milk_yield_messages.append(entry)
        return milk_yield_messages 


    def mod_get_value(self, sheet, row_num, column, exp_type):
        """
           Get the value in a given row and column in excel sheet if it exists.
           :param sheet: a pointer to a sheet of data.
           :param row_num: An int.
           :param column: An int.
           :param exp_type: A Casts enum.
        """
        cell = sheet.cell(row_num, column)
        value = None
        if cell.value == "--" or cell.ctype == xlrd.XL_CELL_EMPTY:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = 0
            elif exp_type == Casts.FLOAT:
                value = 0.0
            elif exp_type == Casts.STR:
                value = ""
            elif exp_type == Casts.BOOL:
                value = None
            else:
                raise ValueError("Unknown type (%s) at row: %d, col: %d!\n" \
                                  % (cell.ctype, row_num, column))
        else:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = int(sheet.cell(row_num, column).value)
            elif exp_type == Casts.STR:
                value = str(sheet.cell(row_num, column).value)
            elif exp_type == Casts.FLOAT:
                value = float(sheet.cell(row_num, column).value)
            else: # Bool
                if sheet.cell(row_num, column).value == "No":
                    value = False
                else:
                    value = True
        return value

    def get_value(self, sheet, row_num, column, exp_type):
        """
           Get the value in a given row and column in excel sheet if it exists.
           :param sheet: a pointer to a sheet of data.
           :param row_num: An int.
           :param column: An int.
           :param exp_type: A Casts enum.
        """
        cell = sheet.cell(row_num, column)
        value = None
        if cell.value == "--" or cell.ctype == xlrd.XL_CELL_EMPTY:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = 0
            elif exp_type == Casts.FLOAT:
                value = 0.0
            elif exp_type == Casts.STR:
                value = ""
            elif exp_type == Casts.BOOL:
                value = None
            else:
                raise ValueError("Unknown type (%s) at row: %d, col: %d!\n" \
                                  % (cell.ctype, row_num, column))
        else:
            if exp_type == Casts.INT or exp_type == Casts.SINT:
                value = int(sheet.cell(row_num, column).value)
            elif exp_type == Casts.STR:
                value = str(sheet.cell(row_num, column).value)
            elif exp_type == Casts.FLOAT:
                value = float(sheet.cell(row_num, column).value)
            else: # Bool
                if sheet.cell(row_num, column).value == "No":
                    value = False
                else:
                    value = True
        return value
