# System packages
from typing import Dict

# Local packages
from sdf.dairymgr.base.global_defs import Casts


# Third party packages
from openpyxl.cell.cell import TYPE_NULL
from openpyxl import load_workbook


__author__ = "Gloire Rubambiza"
__email__ = "gbr26@cornell.edu"
__credits__ = ["Gloire Rubambiza"]


# @brief: A class for reading/manipulating Excel (xlsx) files.
class ExcelReader:


    def __init__(self, file_path, data_only=False):
        """
           Initialize a workbook/worksheet manipulator.
           :param file_path: A string.
           :param data_only: A boolean. Whether to ignore formulas.
        """
        self.file_path = file_path
        self.workbook = load_workbook(self.file_path, data_only=data_only)
        self.active_sheet = self.workbook.active


    def get_cell(self, row, column):
        """
           Get cell value for a given row and column.
           :param row: An int.
           :param column: An int.
        """
        return self.active_sheet.cell(row, column) 


    def switch_worksheet(self, value):
        """
           Switch to another worksheet based on name or index value.
           :param value: Any of str, int.
        """
        self.workbook.active(value)


def get_value(row: Dict,
              column: str,
              filename: str,
              casting: Casts = Casts.NO_CAST
             ):
    
    """
       Get a value from a file if the key:value pair exists.
       This sidesteps some files not having a given key.
       :param row: An object that allows dict-like access.
       :param column: The column name to access.
       :param filename: The filename being access, for logging purposes.
       :param casting: An indicator of any casting to be done.
       :rtype: An int or string or float.
    """
    try:
       if casting == Casts.INT:
           return int(row[column])
       elif casting == Casts.STR:
           return str(row[column])
       elif casting == Casts.FLOAT:
           return float(row[column])
       else: # No casting necessary.
           return row[column]
    except KeyError:
        print("No %s column in %s\n" % (column, filename))
